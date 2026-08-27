# daily_report_bot.py

import pyautogui
import pyperclip
import pygetwindow as gw
import time
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def open_chrome_and_get_price(url):
    """Open Chrome, navigate to the stock page, and copy its visible text."""
    pyautogui.press('win')
    time.sleep(1)
    pyautogui.write('chrome', interval=0.05)
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(3)

    pyautogui.hotkey('ctrl', 'l')
    time.sleep(0.5)
    pyautogui.write(url, interval=0.03)
    pyautogui.press('enter')
    time.sleep(4)

    chrome_windows = gw.getWindowsWithTitle('Chrome')
    if chrome_windows:
        chrome_windows[0].activate()
        time.sleep(1)
    else:
        print("Warning: Chrome window not found!")

    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    return pyperclip.paste()


def extract_stock_value(page_text):
    """Find the first plain price-shaped number on the page."""
    matches = [line.strip() for line in page_text.splitlines()
               if re.fullmatch(r'\d{1,4}\.\d{2}', line.strip())]
    if matches:
        return matches[0]
    return "Data not found"


def type_into_excel(fetched_data, comment):
    """Open Excel as a brand-new blank workbook and type the row (visual demo only, not saved via GUI)."""
    pyautogui.press('win')
    time.sleep(1)
    pyautogui.write('excel', interval=0.05)
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(4)

    excel_windows = gw.getWindowsWithTitle('Excel')
    if excel_windows:
        excel_windows[0].activate()
        time.sleep(1)
    else:
        print("Warning: Excel window not found!")

    pyautogui.hotkey('ctrl', 'n')  # force a brand-new blank workbook — CRITICAL SAFETY FIX
    time.sleep(2)

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    pyautogui.write(now, interval=0.03)
    pyautogui.press('tab')
    pyautogui.write(fetched_data, interval=0.03)
    pyautogui.press('tab')
    pyautogui.write(comment, interval=0.03)
    pyautogui.press('enter')
    time.sleep(1)


def save_report_with_openpyxl(now_str, fetched_data, comment, filepath):
    """Save a neatly formatted .xlsx file with a styled header and bordered cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    headers = ["Date & Time", "Stock Price", "Comment"]
    ws.append(headers)
    ws.append([now_str, fetched_data, comment])

    # --- Header row styling ---
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Borders for all filled cells ---
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            if cell.row == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Column widths so text isn't cramped ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35

    wb.save(filepath)
    print(f"Excel file saved at: {filepath}")


def take_screenshot(filename):
    """Screenshot the current screen (Excel with typed data visible) and save it."""
    screenshot = pyautogui.screenshot()
    screenshot.save(filename)
    print(f"Screenshot saved as {filename}")


def close_apps():
    """Close Chrome and Excel windows opened by the bot. Never saves — always discards changes."""
    # --- Close Chrome ---
    chrome_windows = gw.getWindowsWithTitle('Chrome')
    if chrome_windows:
        chrome_windows[0].activate()
        time.sleep(1)
        pyautogui.hotkey('alt', 'f4')
        time.sleep(1)
        print("Chrome closed.")
    else:
        print("No Chrome window found to close.")

    # --- Close Excel ---
    excel_windows = gw.getWindowsWithTitle('Excel')
    if excel_windows:
        excel_windows[0].activate()
        time.sleep(1)
        pyautogui.hotkey('alt', 'f4')
        time.sleep(2)  # give the "Save changes?" popup time to appear

        # This workbook (Book2) only ever had typed test data —
        # the real data was already saved separately via openpyxl.
        # So we ALWAYS discard here, never save.
        pyautogui.press('n')  # "Don't Save"
        time.sleep(1)
        print("Excel closed (changes discarded — real file already saved via openpyxl).")
    else:
        print("No Excel window found to close.")


def main():
    print("Starting in 5 seconds — don't touch mouse/keyboard...")
    time.sleep(5)

    url = "https://finance.yahoo.com/quote/AAPL"
    page_text = open_chrome_and_get_price(url)
    stock_value = extract_stock_value(page_text)
    print(f"Extracted stock price: {stock_value}")

    comment = "Stock looks stable today"
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    type_into_excel(stock_value, comment)

    # Screenshot BEFORE closing, so we capture the visible typed data
    screenshot_name = f"excel_screenshot_{datetime.now().strftime('%Y-%m-%d')}.png"
    take_screenshot(screenshot_name)

    # Reliably save the real file with openpyxl
    filepath = f"daily_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    save_report_with_openpyxl(now_str, stock_value, comment, filepath)

    # Clean up: close the Chrome and Excel windows the bot opened
    close_apps()


if __name__ == "__main__":
    main()
