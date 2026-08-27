# playwright_assign.py
#
# WhatsApp Web Automation Bot — CAIE Architect Assignment 2

import json
import random
import re
import time
from datetime import datetime

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


def read_contacts(filepath="contacts.xlsx"):
    """Read Name, Phone, Message columns from contacts.xlsx into a list of dicts."""
    wb = load_workbook(filepath)
    ws = wb.active

    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header row
        name, phone, message = row
        if name and phone:
            contacts.append({
                "name": name,
                "phone": normalize_phone(phone),
                "message": message or "Hi {name}, this is an automated message."
            })
    return contacts


def normalize_phone(phone):
    """
    Make sure the phone number is a clean string with a leading '+'.
    Excel often stores a typed '+919812345678' as the plain number 919812345678
    (dropping the '+'), so this puts it back if it's missing.
    """
    phone_str = str(phone).strip()
    # openpyxl can hand back a float like 919812345678.0 for numeric cells
    if phone_str.endswith(".0"):
        phone_str = phone_str[:-2]
    if not phone_str.startswith("+"):
        phone_str = "+" + phone_str
    return phone_str


def wait_for_any_selector(page, selectors, timeout=15000):
    """
    Try a list of possible selectors (WhatsApp Web's internal structure changes
    between updates) and return the first one that appears within `timeout` ms total.
    Raises PlaywrightTimeoutError if none of them show up.
    """
    per_selector_timeout = max(timeout // len(selectors), 3000)
    last_error = None
    for selector in selectors:
        try:
            return page.wait_for_selector(selector, timeout=per_selector_timeout)
        except PlaywrightTimeoutError as e:
            last_error = e
    raise last_error


def human_pause(min_sec=2, max_sec=5):
    """Random delay so the bot doesn't act suspiciously fast/robotic."""
    time.sleep(random.uniform(min_sec, max_sec))


def send_message_to_contact(page, contact):
    """
    Search for the contact, type + send a personalized message, screenshot it.
    Returns a dict describing the outcome.
    """
    name = contact["name"]
    phone = contact["phone"]
    message = contact["message"].replace("{name}", name)

    result = {"name": name, "phone": phone}

    try:
        # 1. Click the search box and type the phone number (most reliable way
        #    to find a contact, since names can be ambiguous / not saved).
        #    WhatsApp Web has changed this element's markup across updates, so
        #    we try a few known variants.
        search_box = wait_for_any_selector(page, [
            'input[aria-label="Search or start a new chat"]',
            'input[data-tab="3"]',
            'div[contenteditable="true"][aria-label="Search input textbox"]',
            'div[contenteditable="true"][data-tab="3"]',
            '#side div[contenteditable="true"]',
        ], timeout=15000)
        search_box.click()
        # Clear any leftover text from a previous contact's search first.
        page.keyboard.press("Control+A")
        page.keyboard.press("Delete")
        human_pause(1, 2)
        search_box.fill(phone)
        human_pause(2, 3)  # give WhatsApp time to actually filter the results

        # 2. Click the result that matches THIS contact's number — not just
        #    "the first row", since that testid is reused for every row in the
        #    whole chat list (including the regular list behind the search
        #    dropdown), which is what caused a message to go to the wrong person.
        # Saved contacts show their NAME in the result row, not their number,
        # so match on either — first name or the last 10 digits of the phone.
        last10 = phone[-10:]
        first_name = name.split()[0] if name else ""
        match_pattern = re.compile(
            f"{re.escape(last10)}|{re.escape(first_name)}", re.IGNORECASE
        )
        try:
            first_result = page.locator(
                'div[data-testid="cell-frame-container"]', has_text=match_pattern
            ).first
            first_result.wait_for(timeout=8000)
            first_result.click()
        except PlaywrightTimeoutError:
            result["status"] = "not_found"
            page.keyboard.press("Escape")  # clear the search box before moving on
            return result

        human_pause(1, 2)

        # SAFETY CHECK: before typing or sending anything, verify the chat that
        # actually opened really belongs to this contact (by checking the phone
        # number appears in the chat header). If it doesn't match, we stop here
        # rather than risk sending to the wrong person.
        try:
            header_text = page.locator('#main header').first.inner_text(timeout=5000)
        except PlaywrightTimeoutError:
            header_text = ""

        phone_matches = last10 in header_text.replace(" ", "").replace("-", "")
        name_matches = name.split()[0].lower() in header_text.lower()
        if not (phone_matches or name_matches):
            result["status"] = "error"
            result["error"] = (
                f"Safety check failed: opened chat header ('{header_text.strip()}') "
                f"doesn't match expected number ({phone}). Message NOT sent."
            )
            page.keyboard.press("Escape")
            return result

        # 3. Click the "Type a message" box at the bottom and type the message.
        message_box = wait_for_any_selector(page, [
            'div[data-testid="conversation-compose-box-input"]',
            'div[contenteditable="true"][aria-placeholder="Type a message"]',
            'div[contenteditable="true"][data-tab="10"]',
            'footer div[contenteditable="true"]',
        ], timeout=10000)
        message_box.click()
        page.keyboard.type(message, delay=random.randint(40, 90))  # human-like typing speed
        human_pause(1, 2)

        # 4. Send it — click the actual Send button rather than relying on a
        #    simulated Enter key, since WhatsApp's newer text editor sometimes
        #    treats a synthetic Enter as "new line" instead of "send".
        #    Confirmed via DevTools: <button data-tab="11" aria-label="Send" ...>
        try:
            send_button = page.wait_for_selector(
                'button[data-tab="11"][aria-label="Send"]', timeout=8000
            )
            send_button.click()
        except PlaywrightTimeoutError:
            page.keyboard.press("Enter")  # fallback if no send button was found

        # 5. Try to confirm the message bubble actually appeared. Best-effort:
        #    WhatsApp's internal class names change often, so if none of these
        #    match we don't fail the whole run — Enter was already pressed above,
        #    which is the actual "send" action.
        try:
            wait_for_any_selector(page, [
                'div[data-testid="msg-container"]',
                'div.message-out',
                'div[class*="message-out"]',
            ], timeout=5000)
        except PlaywrightTimeoutError:
            pass
        human_pause(1, 2)

        # 6. Screenshot the sent message.
        page.screenshot(path=f"sent_{name}.png")

        result["status"] = "sent"

    except Exception as e:
        result["status"] = "error"
        result["error"] = str(e)

    return result


def extract_last_messages(page, contact, count=3):
    """
    With the contact's chat open, return the text of the last `count` messages.
    Scoped to the messages panel only, so it can never accidentally pick up
    whatever is currently (un-sent) typed in the compose box at the bottom.
    """
    container_selectors = [
        'div[data-testid="conversation-panel-messages"]',
    ]
    container_sel = None
    for sel in container_selectors:
        try:
            page.wait_for_selector(sel, timeout=4000)
            container_sel = sel
            break
        except PlaywrightTimeoutError:
            continue

    if not container_sel:
        return []

    for msg_sel in ['span.selectable-text', 'div.selectable-text', 'div.copyable-text']:
        bubbles = page.locator(f"{container_sel} {msg_sel}").all_inner_texts()
        if bubbles:
            return bubbles[-count:]
    return []


def save_reports(results, out_dir="."):
    """Save the full run results as both JSON and a summary Excel file, dated today."""
    today = datetime.now().strftime("%Y-%m-%d")

    json_path = f"{out_dir}/whatsapp_report_{today}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"JSON report saved: {json_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "WhatsApp Report"
    ws.append(["Name", "Phone", "Status", "Last Messages"])

    for r in results:
        last_msgs = " | ".join(r.get("last_messages", []))
        ws.append([r.get("name"), r.get("phone"), r.get("status"), last_msgs])

    xlsx_path = f"{out_dir}/whatsapp_report_{today}.xlsx"
    wb.save(xlsx_path)
    print(f"Excel report saved: {xlsx_path}")


def main():
    contacts = read_contacts("contacts.xlsx")
    print(f"Loaded {len(contacts)} contacts.")

    results = []

    with sync_playwright() as p:
        # A persistent context saves your WhatsApp Web login to the
        # "whatsapp_session" folder, so after the very first run you won't
        # need to scan the QR code again.
        context = p.chromium.launch_persistent_context(
            "whatsapp_session", headless=False
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://web.whatsapp.com")

        print("Waiting for WhatsApp Web to load (scan the QR code now if this is the first run)...")
        # Automatically wait for the search box to appear — no keypress needed.
        # This polls for up to 90 seconds, which is plenty of time to scan the
        # QR code by hand on a first run, and is nearly instant on later runs.
        page.wait_for_selector(
            'input[aria-label="Search or start a new chat"]', timeout=90000
        )
        print("Logged in. Starting to send messages...")

        for contact in contacts:
            result = send_message_to_contact(page, contact)

            if result.get("status") == "sent":
                result["last_messages"] = extract_last_messages(page, contact)

            print(f"{contact['name']}: {result['status']}")
            results.append(result)

            human_pause()  # human-like gap before moving to the next contact

        context.close()

    save_reports(results)
    print("Done. Reports saved.")


if __name__ == "__main__":
    main()
